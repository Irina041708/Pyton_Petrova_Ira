from tkinter import Button, Entry, Label, Tk, messagebox
from tkinter import *
import xxml
import openpyxl
from openpyxl import load_workbook


data_xlsx = "data_file.xlsx"
wb = load_workbook(data_xlsx)
ws = wb['data_list']
new_row1 = input("Введите фамилию: ")
new_row2 = input("Введите имя: ")
new_row3 = input("Введите год рождения: ")
new_row4 = input("Введите телефон: ")
ws.append([new_row1,new_row2,new_row3,new_row4])

wb.save("data__file.xlsx")
wb.close()


def get_add_user_xlsx():
        data_xlsx = "data_file.xlsx"
        wb = load_workbook(data_xlsx)
        ws = wb['data_list']
        
        data_list = (e.get(),e.get(),e.get(),lb['text'])
        ws.append(data_list)
        wb.save(data_xlsx)
        wb.close()
        messagebox.askokcancel('Сохранение','Успешно сохранено')
       
root = Tk()
root.title('Добавление строки')
root.geometry('600x200')
root.resizable(0,0)

e = Entry(root)
e.pack()
lb = Label(root,text = 'Фамилия' , font='Arial 15 bold')
lb.pack()
btn = Button(root, text='Сохранить', font='Arial 15 bold', command=get_add_user_xlsx)
btn.pack()

root.mainloop()
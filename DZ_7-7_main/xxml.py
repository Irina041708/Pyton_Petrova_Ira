import openpyxl
from openpyxl import load_workbook
import pandas as pd 


def read_file():
    wb = openpyxl.load_workbook('data_file.xlsx')
    data_xlsx = wb['data_list']
    wb.close()
    return data_xlsx


def read_file_2():
    data_xlsx_2 = pd.read_excel('data_file.xlsx')
    return data_xlsx_2


# def write_in_file(data_xlsx):
#     wb = load_workbook("data_file.xlsx")
#     ws = wb.active
#     for row in ws.rows:
#         data_xlsx1 = row[0]
#         data_xlsx2 = row[1]
#         data_xlsx3 = row[2]
#         data_xlsx4 = row[3]

#         Фамилия = data_xlsx.value
#         Имя = data_xlsx.value
#         Год_рождения = data_xlsx.value
#         Телефон = data_xlsx.value
#         print(Фамилия, Имя, Год_рождения, Телефон)
#     ws.append([Фамилия, Имя,Год_рождения,Телефон])
#     wb.save("data_file.xlsx")
#     wb.close()
